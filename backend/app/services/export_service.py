"""Excel export service for session history and admin reports."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.session import CoachingSession
from app.services.analytics_service import get_org_analytics


async def export_sessions_excel(db: AsyncSession, user_id: str) -> BytesIO:
    """Generate Excel workbook with user's session history."""
    result = await db.execute(
        select(CoachingSession)
        .options(selectinload(CoachingSession.scenario))
        .where(
            and_(
                CoachingSession.user_id == user_id,
                CoachingSession.status == "scored",
            )
        )
        .order_by(CoachingSession.completed_at.desc())
    )
    sessions = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Session History"

    # Header row
    headers = ["Date", "Scenario", "Score", "Passed", "Duration (min)", "Session Type"]
    header_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, session in enumerate(sessions, 2):
        ws.cell(
            row=row_idx,
            column=1,
            value=session.completed_at.strftime("%Y-%m-%d") if session.completed_at else "",
        )
        ws.cell(
            row=row_idx,
            column=2,
            value=session.scenario.name if session.scenario else "",
        )
        ws.cell(row=row_idx, column=3, value=session.overall_score)
        ws.cell(row=row_idx, column=4, value="Yes" if session.passed else "No")
        ws.cell(
            row=row_idx,
            column=5,
            value=session.duration_seconds // 60 if session.duration_seconds else "",
        )
        ws.cell(row=row_idx, column=6, value=session.session_type)

    # Column widths
    widths = [15, 30, 10, 10, 15, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def export_admin_report_excel(db: AsyncSession) -> BytesIO:
    """Generate Excel workbook with org-level admin report."""
    org = await get_org_analytics(db)

    wb = Workbook()
    header_font = Font(bold=True)

    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Overview"
    overview_data = [
        ("Metric", "Value"),
        ("Total Users", org.total_users),
        ("Active Users", org.active_users),
        ("Completion Rate (%)", org.completion_rate),
        ("Total Sessions", org.total_sessions),
        ("Average Score", org.avg_org_score),
    ]
    for row_idx, (label, value) in enumerate(overview_data, 1):
        cell_a = ws_overview.cell(row=row_idx, column=1, value=label)
        ws_overview.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = header_font
            ws_overview.cell(row=row_idx, column=2).font = header_font
    ws_overview.column_dimensions["A"].width = 25
    ws_overview.column_dimensions["B"].width = 15

    # Sheet 2: BU Comparison
    ws_bu = wb.create_sheet("BU Comparison")
    bu_headers = ["Business Unit", "Sessions", "Avg Score", "Users"]
    for col, header in enumerate(bu_headers, 1):
        cell = ws_bu.cell(row=1, column=col, value=header)
        cell.font = header_font
    for row_idx, bu in enumerate(org.bu_stats, 2):
        ws_bu.cell(row=row_idx, column=1, value=bu.business_unit)
        ws_bu.cell(row=row_idx, column=2, value=bu.session_count)
        ws_bu.cell(row=row_idx, column=3, value=bu.avg_score)
        ws_bu.cell(row=row_idx, column=4, value=bu.user_count)
    for col, width in enumerate([25, 12, 12, 10], 1):
        ws_bu.column_dimensions[ws_bu.cell(row=1, column=col).column_letter].width = width

    # Sheet 3: Skill Gaps
    ws_gaps = wb.create_sheet("Skill Gaps")
    if org.skill_gaps:
        # Collect unique dimensions and BUs
        dimensions = sorted({cell.dimension for cell in org.skill_gaps})
        bus = sorted({cell.business_unit for cell in org.skill_gaps})
        gap_lookup = {
            (cell.business_unit, cell.dimension): cell.avg_score for cell in org.skill_gaps
        }

        # Header row
        ws_gaps.cell(row=1, column=1, value="Business Unit").font = header_font
        for col, dim in enumerate(dimensions, 2):
            ws_gaps.cell(row=1, column=col, value=dim).font = header_font

        # Data rows
        for row_idx, bu in enumerate(bus, 2):
            ws_gaps.cell(row=row_idx, column=1, value=bu)
            for col, dim in enumerate(dimensions, 2):
                ws_gaps.cell(row=row_idx, column=col, value=gap_lookup.get((bu, dim), ""))
        ws_gaps.column_dimensions["A"].width = 25

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
