"""Text extraction from training material documents (PDF, DOCX, XLSX)."""

import io

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

CHUNK_SIZE = 2000
CHUNK_OVERLAP = 200


def extract_text(content: bytes, content_type: str) -> list[tuple[str, str]]:
    """Extract text from document bytes.

    Returns list of (page_label, text_content) tuples.
    Each tuple represents one chunk suitable for RAG indexing.
    """
    if content_type == "application/pdf":
        return _extract_pdf(content)
    elif content_type in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ):
        return _extract_docx(content)
    elif content_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",):
        return _extract_xlsx(content)
    return []


def _extract_pdf(content: bytes) -> list[tuple[str, str]]:
    """Extract text page-by-page from PDF."""
    reader = PdfReader(io.BytesIO(content))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append((f"Page {i + 1}", text))
    return pages


def _extract_docx(content: bytes) -> list[tuple[str, str]]:
    """Extract text from DOCX, chunked by paragraph groups."""
    doc = Document(io.BytesIO(content))
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if not full_text.strip():
        return []
    return _chunk_text(full_text, prefix="Section")


def _extract_xlsx(content: bytes) -> list[tuple[str, str]]:
    """Extract text from XLSX, one chunk per sheet."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    pages = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            pages.append((f"Sheet: {sheet_name}", text))
    wb.close()
    return pages


def _chunk_text(
    text: str,
    prefix: str = "Chunk",
    chunk_size: int = CHUNK_SIZE,
    overlap: int = CHUNK_OVERLAP,
) -> list[tuple[str, str]]:
    """Split text into overlapping chunks of approximately chunk_size characters."""
    if len(text) <= chunk_size:
        return [(f"{prefix} 1", text)]

    chunks = []
    start = 0
    chunk_num = 1
    while start < len(text):
        end = start + chunk_size
        # Try to break at a newline near the boundary
        if end < len(text):
            newline_pos = text.rfind("\n", start + chunk_size - overlap, end + overlap)
            if newline_pos > start:
                end = newline_pos + 1
        chunk = text[start:end].strip()
        if chunk:
            chunks.append((f"{prefix} {chunk_num}", chunk))
            chunk_num += 1
        start = end - overlap
        if start >= len(text):
            break
    return chunks
