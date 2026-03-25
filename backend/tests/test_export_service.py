"""Unit tests for export_service: Excel generation for sessions and admin reports."""

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.models.hcp_profile import HcpProfile
from app.models.scenario import Scenario
from app.models.score import ScoreDetail, SessionScore
from app.models.session import CoachingSession
from app.models.user import User
from app.services.auth import get_password_hash
from app.services.export_service import export_admin_report_excel, export_sessions_excel


async def _create_user(session, *, username="user1", role="user", bu="BU-Sales") -> User:
    """Helper to create a user in the test DB."""
    user = User(
        username=username,
        email=f"{username}@test.com",
        hashed_password=get_password_hash("pass"),
        full_name=f"Test {username}",
        role=role,
        business_unit=bu,
    )
    session.add(user)
    await session.flush()
    return user


async def _create_hcp(session, *, created_by: str) -> HcpProfile:
    """Helper to create an HCP profile."""
    hcp = HcpProfile(
        name="Dr. Export",
        specialty="Oncology",
        created_by=created_by,
    )
    session.add(hcp)
    await session.flush()
    return hcp


async def _create_scenario(session, *, hcp_id: str, created_by: str, name: str = "Test Scenario"):
    """Helper to create a scenario."""
    scenario = Scenario(
        name=name,
        product="TestDrug",
        difficulty="medium",
        status="active",
        hcp_profile_id=hcp_id,
        created_by=created_by,
    )
    session.add(scenario)
    await session.flush()
    return scenario


async def _create_scored_session(
    session,
    *,
    user_id: str,
    scenario_id: str,
    overall_score: float = 80.0,
    passed: bool = True,
    completed_at: datetime | None = None,
    duration_seconds: int = 600,
    session_type: str = "f2f",
) -> CoachingSession:
    """Helper to create a scored coaching session."""
    cs = CoachingSession(
        user_id=user_id,
        scenario_id=scenario_id,
        status="scored",
        overall_score=overall_score,
        passed=passed,
        completed_at=completed_at or datetime.now(UTC),
        duration_seconds=duration_seconds,
        session_type=session_type,
    )
    session.add(cs)
    await session.flush()
    return cs


# ---------------------------------------------------------------------------
# export_sessions_excel
# ---------------------------------------------------------------------------


class TestExportSessionsExcel:
    """Tests for export_sessions_excel."""

    async def test_returns_bytesio(self, db_session):
        """Should return a BytesIO object."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        assert isinstance(result, BytesIO)

    async def test_workbook_has_correct_sheet_name(self, db_session):
        """Sheet should be named 'Session History'."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        assert wb.sheetnames == ["Session History"]
        wb.close()

    async def test_header_row(self, db_session):
        """First row should contain expected headers."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["Date", "Scenario", "Score", "Passed", "Duration (min)", "Session Type"]
        wb.close()

    async def test_header_font_is_bold(self, db_session):
        """Header cells should be bold."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        for col in range(1, 7):
            assert ws.cell(row=1, column=col).font.bold is True
        wb.close()

    async def test_data_rows_populated(self, db_session):
        """Data rows should match sessions in the database."""
        user = await _create_user(db_session)
        hcp = await _create_hcp(db_session, created_by=user.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=user.id)
        completed = datetime(2025, 6, 15, 10, 0, 0, tzinfo=UTC)
        await _create_scored_session(
            db_session,
            user_id=user.id,
            scenario_id=scenario.id,
            overall_score=85.0,
            passed=True,
            completed_at=completed,
            duration_seconds=720,
            session_type="f2f",
        )
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "2025-06-15"
        assert ws.cell(row=2, column=2).value == "Test Scenario"
        assert ws.cell(row=2, column=3).value == 85.0
        assert ws.cell(row=2, column=4).value == "Yes"
        assert ws.cell(row=2, column=5).value == 12  # 720 // 60
        assert ws.cell(row=2, column=6).value == "f2f"
        wb.close()

    async def test_empty_sessions(self, db_session):
        """User with no scored sessions should produce header-only workbook."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        # Header row exists
        assert ws.cell(row=1, column=1).value == "Date"
        # No data rows
        assert ws.cell(row=2, column=1).value is None
        wb.close()

    async def test_column_widths_set(self, db_session):
        """Column widths should match expected values."""
        user = await _create_user(db_session)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        expected_widths = {"A": 15, "B": 30, "C": 10, "D": 10, "E": 15, "F": 15}
        for letter, expected in expected_widths.items():
            assert ws.column_dimensions[letter].width == expected
        wb.close()

    async def test_failed_session_shows_no(self, db_session):
        """A session that did not pass should show 'No'."""
        user = await _create_user(db_session)
        hcp = await _create_hcp(db_session, created_by=user.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=user.id)
        await _create_scored_session(
            db_session,
            user_id=user.id,
            scenario_id=scenario.id,
            overall_score=55.0,
            passed=False,
        )
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(row=2, column=4).value == "No"
        wb.close()

    async def test_null_completed_at_shows_empty(self, db_session):
        """Session with no completed_at should show empty date."""
        user = await _create_user(db_session)
        hcp = await _create_hcp(db_session, created_by=user.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=user.id)
        cs = CoachingSession(
            user_id=user.id,
            scenario_id=scenario.id,
            status="scored",
            overall_score=75.0,
            passed=True,
            completed_at=None,
            duration_seconds=None,
        )
        db_session.add(cs)
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        # Empty string written by openpyxl loads back as None
        assert ws.cell(row=2, column=1).value in ("", None)
        assert ws.cell(row=2, column=5).value in ("", None)
        wb.close()

    async def test_multiple_sessions_ordered_desc(self, db_session):
        """Sessions should be ordered by completed_at descending."""
        user = await _create_user(db_session)
        hcp = await _create_hcp(db_session, created_by=user.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=user.id)
        await _create_scored_session(
            db_session,
            user_id=user.id,
            scenario_id=scenario.id,
            overall_score=70.0,
            completed_at=datetime(2025, 1, 1, tzinfo=UTC),
        )
        await _create_scored_session(
            db_session,
            user_id=user.id,
            scenario_id=scenario.id,
            overall_score=90.0,
            completed_at=datetime(2025, 6, 1, tzinfo=UTC),
        )
        await db_session.commit()

        result = await export_sessions_excel(db_session, user.id)
        wb = load_workbook(result)
        ws = wb.active
        # Most recent first
        assert ws.cell(row=2, column=3).value == 90.0
        assert ws.cell(row=3, column=3).value == 70.0
        wb.close()


# ---------------------------------------------------------------------------
# export_admin_report_excel
# ---------------------------------------------------------------------------


class TestExportAdminReportExcel:
    """Tests for export_admin_report_excel."""

    async def test_returns_bytesio(self, db_session):
        """Should return a BytesIO object."""
        result = await export_admin_report_excel(db_session)
        assert isinstance(result, BytesIO)

    async def test_has_three_sheets(self, db_session):
        """Workbook should have Overview, BU Comparison, and Skill Gaps sheets."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        assert wb.sheetnames == ["Overview", "BU Comparison", "Skill Gaps"]
        wb.close()

    async def test_overview_sheet_content(self, db_session):
        """Overview sheet should have Metric/Value header and org stats."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Overview"]
        assert ws.cell(row=1, column=1).value == "Metric"
        assert ws.cell(row=1, column=2).value == "Value"
        assert ws.cell(row=2, column=1).value == "Total Users"
        assert ws.cell(row=3, column=1).value == "Active Users"
        assert ws.cell(row=4, column=1).value == "Completion Rate (%)"
        assert ws.cell(row=5, column=1).value == "Total Sessions"
        assert ws.cell(row=6, column=1).value == "Average Score"
        wb.close()

    async def test_overview_header_bold(self, db_session):
        """Header row in Overview should be bold."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Overview"]
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=2).font.bold is True
        wb.close()

    async def test_overview_column_widths(self, db_session):
        """Overview sheet should have correct column widths."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Overview"]
        assert ws.column_dimensions["A"].width == 25
        assert ws.column_dimensions["B"].width == 15
        wb.close()

    async def test_bu_comparison_headers(self, db_session):
        """BU Comparison sheet should have correct headers."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["BU Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["Business Unit", "Sessions", "Avg Score", "Users"]
        wb.close()

    async def test_bu_comparison_data(self, db_session):
        """BU Comparison sheet should contain BU stats data."""
        admin = await _create_user(db_session, username="admin", role="admin", bu="HQ")
        user1 = await _create_user(db_session, username="user1", role="user", bu="BU-Sales")
        hcp = await _create_hcp(db_session, created_by=admin.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=admin.id)
        cs = await _create_scored_session(
            db_session, user_id=user1.id, scenario_id=scenario.id, overall_score=80.0
        )
        score = SessionScore(session_id=cs.id, overall_score=80.0, passed=True)
        db_session.add(score)
        await db_session.flush()
        detail = ScoreDetail(score_id=score.id, dimension="key_message", score=80.0, weight=30)
        db_session.add(detail)
        await db_session.commit()

        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["BU Comparison"]
        assert ws.cell(row=2, column=1).value == "BU-Sales"
        assert ws.cell(row=2, column=2).value == 1  # session count
        wb.close()

    async def test_skill_gaps_sheet_with_data(self, db_session):
        """Skill Gaps sheet should have BU rows and dimension columns."""
        admin = await _create_user(db_session, username="admin", role="admin", bu="HQ")
        user1 = await _create_user(db_session, username="user1", role="user", bu="BU-Sales")
        hcp = await _create_hcp(db_session, created_by=admin.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=admin.id)
        cs = await _create_scored_session(
            db_session, user_id=user1.id, scenario_id=scenario.id, overall_score=80.0
        )
        score = SessionScore(session_id=cs.id, overall_score=80.0, passed=True)
        db_session.add(score)
        await db_session.flush()
        detail = ScoreDetail(score_id=score.id, dimension="key_message", score=85.0, weight=30)
        db_session.add(detail)
        await db_session.commit()

        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Skill Gaps"]
        # Header row
        assert ws.cell(row=1, column=1).value == "Business Unit"
        assert ws.cell(row=1, column=2).value == "key_message"
        # Data row
        assert ws.cell(row=2, column=1).value == "BU-Sales"
        assert ws.cell(row=2, column=2).value == 85.0
        wb.close()

    async def test_skill_gaps_empty(self, db_session):
        """Skill Gaps sheet with no data should be empty."""
        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Skill Gaps"]
        # No header beyond the sheet itself since skill_gaps is empty
        assert ws.cell(row=1, column=1).value is None
        wb.close()

    async def test_overview_values_match_org_analytics(self, db_session):
        """Overview sheet values should match computed org analytics."""
        admin = await _create_user(db_session, username="admin", role="admin", bu="HQ")
        user1 = await _create_user(db_session, username="user1", role="user", bu="BU-A")
        user2 = await _create_user(db_session, username="user2", role="user", bu="BU-A")
        hcp = await _create_hcp(db_session, created_by=admin.id)
        scenario = await _create_scenario(db_session, hcp_id=hcp.id, created_by=admin.id)
        await _create_scored_session(
            db_session, user_id=user1.id, scenario_id=scenario.id, overall_score=80.0
        )
        await _create_scored_session(
            db_session, user_id=user2.id, scenario_id=scenario.id, overall_score=90.0
        )
        await db_session.commit()

        result = await export_admin_report_excel(db_session)
        wb = load_workbook(result)
        ws = wb["Overview"]
        assert ws.cell(row=2, column=2).value == 2  # total users
        assert ws.cell(row=3, column=2).value == 2  # active users
        assert ws.cell(row=4, column=2).value == 100.0  # completion rate
        assert ws.cell(row=5, column=2).value == 2  # total sessions
        assert ws.cell(row=6, column=2).value == 85.0  # avg score
        wb.close()
